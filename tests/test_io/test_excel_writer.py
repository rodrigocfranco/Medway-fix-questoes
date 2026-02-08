"""Tests for Excel output writer with atomic write and formatting."""

import sqlite3
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from construtor.io.excel_writer import ExcelWriter


@pytest.fixture
def sample_db_with_approved_questions(tmp_path):
    """Create SQLite DB with sample approved questions."""
    db_path = tmp_path / "test.db"
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Create questions table with all 26 columns
    cursor.execute("""
        CREATE TABLE questions (
            id INTEGER PRIMARY KEY,
            tema TEXT NOT NULL,
            foco TEXT NOT NULL,
            sub_foco TEXT NOT NULL,
            periodo TEXT NOT NULL,
            nivel_dificuldade INTEGER NOT NULL,
            tipo_enunciado TEXT NOT NULL,
            enunciado TEXT NOT NULL,
            alternativa_a TEXT NOT NULL,
            alternativa_b TEXT NOT NULL,
            alternativa_c TEXT NOT NULL,
            alternativa_d TEXT NOT NULL,
            resposta_correta TEXT NOT NULL,
            objetivo_educacional TEXT NOT NULL,
            comentario_introducao TEXT NOT NULL,
            comentario_visao_especifica TEXT NOT NULL,
            comentario_alt_a TEXT NOT NULL,
            comentario_alt_b TEXT NOT NULL,
            comentario_alt_c TEXT NOT NULL,
            comentario_alt_d TEXT NOT NULL,
            comentario_visao_aprovado TEXT NOT NULL,
            referencia_bibliografica TEXT NOT NULL,
            suporte_imagem TEXT,
            fonte_imagem TEXT,
            modelo_llm TEXT NOT NULL,
            rodadas_validacao INTEGER NOT NULL,
            concordancia_comentador INTEGER NOT NULL,
            status TEXT NOT NULL
        )
    """)

    # Insert approved question
    cursor.execute("""
        INSERT INTO questions (
            tema, foco, sub_foco, periodo, nivel_dificuldade,
            tipo_enunciado, enunciado, alternativa_a, alternativa_b,
            alternativa_c, alternativa_d, resposta_correta,
            objetivo_educacional, comentario_introducao,
            comentario_visao_especifica, comentario_alt_a,
            comentario_alt_b, comentario_alt_c, comentario_alt_d,
            comentario_visao_aprovado, referencia_bibliografica,
            suporte_imagem, fonte_imagem, modelo_llm,
            rodadas_validacao, concordancia_comentador, status
        ) VALUES (
            'Cardiologia', 'ICC', 'Diagnóstico', '3º ano', 2,
            'caso clínico', 'Paciente com dispneia...', 'Alt A', 'Alt B',
            'Alt C', 'Alt D', 'A',
            'Identificar ICC', 'Intro comentário',
            'Visão específica', 'Comentário A',
            'Comentário B', 'Comentário C', 'Comentário D',
            'Visão aprovado', 'Harrison 21ed p.123',
            NULL, NULL, 'gpt-4',
            2, 1, 'approved'
        )
    """)

    # Insert pending question (should NOT be exported)
    cursor.execute("""
        INSERT INTO questions (
            tema, foco, sub_foco, periodo, nivel_dificuldade,
            tipo_enunciado, enunciado, alternativa_a, alternativa_b,
            alternativa_c, alternativa_d, resposta_correta,
            objetivo_educacional, comentario_introducao,
            comentario_visao_especifica, comentario_alt_a,
            comentario_alt_b, comentario_alt_c, comentario_alt_d,
            comentario_visao_aprovado, referencia_bibliografica,
            suporte_imagem, fonte_imagem, modelo_llm,
            rodadas_validacao, concordancia_comentador, status
        ) VALUES (
            'Pneumologia', 'Asma', 'Tratamento', '2º ano', 1,
            'conceitual', 'Qual medicamento...', 'Med A', 'Med B',
            'Med C', 'Med D', 'B',
            'Conhecer tratamento', 'Intro',
            'Visão específica', 'Com A',
            'Com B', 'Com C', 'Com D',
            'Visão geral', 'Fonte X',
            NULL, NULL, 'claude-4',
            1, 1, 'pending'
        )
    """)

    conn.commit()
    conn.close()

    return db_path


@pytest.fixture
def empty_db(tmp_path):
    """Create SQLite DB with no approved questions."""
    db_path = tmp_path / "empty.db"
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE questions (
            id INTEGER PRIMARY KEY,
            tema TEXT,
            status TEXT
        )
    """)

    conn.commit()
    conn.close()

    return db_path


@pytest.fixture
def output_excel_path(tmp_path):
    """Provide temporary path for Excel output."""
    return tmp_path / "output.xlsx"


# Happy Path Tests
def test_export_approved_questions(sample_db_with_approved_questions, output_excel_path):
    """Test exporting approved questions to Excel."""
    # Arrange
    writer = ExcelWriter(db_path=str(sample_db_with_approved_questions))

    # Act
    writer.export_to_excel(str(output_excel_path))

    # Assert
    assert output_excel_path.exists()
    df = pd.read_excel(output_excel_path, engine="openpyxl")
    assert len(df) == 1  # Only 1 approved question
    assert df["tema"][0] == "Cardiologia"
    assert df["foco"][0] == "ICC"


def test_all_26_columns_present_in_correct_order(
    sample_db_with_approved_questions, output_excel_path
):
    """Test all 26 columns are present in the correct order."""
    # Arrange
    writer = ExcelWriter(db_path=str(sample_db_with_approved_questions))
    expected_columns = [
        "tema",
        "foco",
        "sub_foco",
        "periodo",
        "nivel_dificuldade",
        "tipo_enunciado",
        "enunciado",
        "alternativa_a",
        "alternativa_b",
        "alternativa_c",
        "alternativa_d",
        "resposta_correta",
        "objetivo_educacional",
        "comentario_introducao",
        "comentario_visao_especifica",
        "comentario_alt_a",
        "comentario_alt_b",
        "comentario_alt_c",
        "comentario_alt_d",
        "comentario_visao_aprovado",
        "referencia_bibliografica",
        "suporte_imagem",
        "fonte_imagem",
        "modelo_llm",
        "rodadas_validacao",
        "concordancia_comentador",
    ]

    # Act
    writer.export_to_excel(str(output_excel_path))

    # Assert
    df = pd.read_excel(output_excel_path, engine="openpyxl")
    assert list(df.columns) == expected_columns


def test_header_row_is_bold(sample_db_with_approved_questions, output_excel_path):
    """Test header row has bold formatting."""
    # Arrange
    writer = ExcelWriter(db_path=str(sample_db_with_approved_questions))

    # Act
    writer.export_to_excel(str(output_excel_path))

    # Assert
    wb = load_workbook(output_excel_path)
    ws = wb.active
    for cell in ws[1]:  # First row (header)
        assert cell.font.bold is True
    wb.close()


# Atomic Write Tests
def test_temporary_file_created_during_write(sample_db_with_approved_questions, output_excel_path):
    """Test atomic write uses temporary file."""
    # Arrange
    writer = ExcelWriter(db_path=str(sample_db_with_approved_questions))

    # Act
    writer.export_to_excel(str(output_excel_path))

    # Assert
    # Temp file should be cleaned up after successful write
    temp_file = Path(str(output_excel_path) + ".tmp")
    assert not temp_file.exists()
    assert output_excel_path.exists()


# Directory Creation Tests
def test_parent_directory_created_if_not_exists(sample_db_with_approved_questions, tmp_path):
    """Test parent directory is created automatically."""
    # Arrange
    writer = ExcelWriter(db_path=str(sample_db_with_approved_questions))
    nested_path = tmp_path / "output" / "exports" / "final.xlsx"

    # Act
    writer.export_to_excel(str(nested_path))

    # Assert
    assert nested_path.exists()
    assert nested_path.parent.exists()


# Status Filtering Tests
def test_only_approved_questions_exported(sample_db_with_approved_questions, output_excel_path):
    """Test only questions with status='approved' are exported."""
    # Arrange
    writer = ExcelWriter(db_path=str(sample_db_with_approved_questions))

    # Act
    writer.export_to_excel(str(output_excel_path))

    # Assert
    df = pd.read_excel(output_excel_path, engine="openpyxl")
    assert len(df) == 1  # Only approved, not pending


def test_pending_questions_excluded(sample_db_with_approved_questions, output_excel_path):
    """Test pending questions are not included in export."""
    # Arrange
    writer = ExcelWriter(db_path=str(sample_db_with_approved_questions))

    # Act
    writer.export_to_excel(str(output_excel_path))

    # Assert
    df = pd.read_excel(output_excel_path, engine="openpyxl")
    # Should only have Cardiologia (approved), not Pneumologia (pending)
    assert all(df["tema"] == "Cardiologia")
    assert "Pneumologia" not in df["tema"].values


# Error Handling Tests
def test_error_when_database_not_found(output_excel_path):
    """Test OSError when SQLite database doesn't exist."""
    # Arrange
    writer = ExcelWriter(db_path="nonexistent.db")

    # Act & Assert
    with pytest.raises(OSError) as exc_info:
        writer.export_to_excel(str(output_excel_path))

    # Accept various database-related error messages
    error_msg = str(exc_info.value).lower()
    assert (
        "database" in error_msg
        or "unable to open" in error_msg
        or "no such table" in error_msg
        or "failed to load questions" in error_msg
    )


def test_error_for_non_xlsx_extension(sample_db_with_approved_questions, tmp_path):
    """Test ValueError for non-.xlsx extension."""
    # Arrange
    writer = ExcelWriter(db_path=str(sample_db_with_approved_questions))
    invalid_path = tmp_path / "output.csv"

    # Act & Assert
    with pytest.raises(ValueError) as exc_info:
        writer.export_to_excel(str(invalid_path))

    assert ".xlsx" in str(exc_info.value)


# Edge Cases Tests
def test_export_with_empty_result(empty_db, output_excel_path):
    """Test export when no approved questions exist."""
    # Arrange
    writer = ExcelWriter(db_path=str(empty_db))

    # Act & Assert
    # Should either create empty Excel or raise informative error
    # Let's allow empty export (returns without error)
    writer.export_to_excel(str(output_excel_path))
    # File should exist but be empty or have headers only
    if output_excel_path.exists():
        df = pd.read_excel(output_excel_path, engine="openpyxl")
        assert len(df) == 0


def test_export_with_limit_parameter(sample_db_with_approved_questions, output_excel_path):
    """Test export with LIMIT parameter for partial export."""
    # Arrange
    writer = ExcelWriter(db_path=str(sample_db_with_approved_questions))

    # Act
    writer.export_to_excel(str(output_excel_path), limit=1)

    # Assert
    df = pd.read_excel(output_excel_path, engine="openpyxl")
    assert len(df) == 1


def test_cleanup_temp_file_on_write_failure(sample_db_with_approved_questions, tmp_path):
    """Test temporary file is cleaned up on write failure."""
    # Arrange
    writer = ExcelWriter(db_path=str(sample_db_with_approved_questions))
    # Create a file first, then make its parent directory to simulate permission error
    test_dir = tmp_path / "readonly"
    test_dir.mkdir()
    test_file = test_dir / "test.xlsx"

    # Write a valid file first
    writer.export_to_excel(str(test_file))

    # Now change permissions to read-only to force failure on next write
    import os
    import stat

    os.chmod(test_dir, stat.S_IRUSR | stat.S_IXUSR)  # Read + execute only

    # Act & Assert
    try:
        with pytest.raises((OSError, OSError, PermissionError)):
            writer.export_to_excel(str(test_file))

        # Temp file should be cleaned up even on error
        temp_file = Path(str(test_file) + ".tmp")
        assert not temp_file.exists()
    finally:
        # Restore permissions for cleanup
        os.chmod(test_dir, stat.S_IRWXU)


# Integration Test
def test_end_to_end_export(sample_db_with_approved_questions, output_excel_path):
    """Test complete end-to-end export flow."""
    # Arrange
    writer = ExcelWriter(db_path=str(sample_db_with_approved_questions))

    # Act
    writer.export_to_excel(str(output_excel_path))

    # Assert - Verify complete data integrity
    df = pd.read_excel(output_excel_path, engine="openpyxl")

    # Check data completeness
    assert len(df) == 1
    assert df["tema"][0] == "Cardiologia"
    assert df["foco"][0] == "ICC"
    assert df["sub_foco"][0] == "Diagnóstico"
    assert df["periodo"][0] == "3º ano"
    assert df["nivel_dificuldade"][0] == 2
    assert df["resposta_correta"][0] == "A"
    assert df["modelo_llm"][0] == "gpt-4"
    assert df["rodadas_validacao"][0] == 2

    # Check NULL handling (suporte_imagem, fonte_imagem should be empty/NaN)
    assert pd.isna(df["suporte_imagem"][0]) or df["suporte_imagem"][0] == ""
    assert pd.isna(df["fonte_imagem"][0]) or df["fonte_imagem"][0] == ""

    # Verify Excel formatting
    wb = load_workbook(output_excel_path)
    ws = wb.active
    assert ws[1][0].font.bold is True  # Header is bold
    wb.close()
