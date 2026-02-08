"""Excel output file writer with atomic write and formatting."""

import logging
import os
import sqlite3
from pathlib import Path
from typing import ClassVar

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Excel output file writer with atomic write and formatting.

    Exports approved questions from SQLite to Excel format with:
    - All 26 columns in correct order
    - Atomic write pattern (temp file + rename) to prevent corruption
    - Bold header formatting
    - Automatic directory creation
    """

    COLUMN_ORDER: ClassVar[list[str]] = [
        "tema",
        "foco",
        "sub_foco",
        "periodo",
        "nivel_dificuldade",
        "tipo_enunciado",
        "enunciado",
        "alternativa_a",
        "alternativa_b",
        "alternativa_c",
        "alternativa_d",
        "resposta_correta",
        "objetivo_educacional",
        "comentario_introducao",
        "comentario_visao_especifica",
        "comentario_alt_a",
        "comentario_alt_b",
        "comentario_alt_c",
        "comentario_alt_d",
        "comentario_visao_aprovado",
        "referencia_bibliografica",
        "suporte_imagem",
        "fonte_imagem",
        "modelo_llm",
        "rodadas_validacao",
        "concordancia_comentador",
    ]

    DEFAULT_SHEET_NAME: ClassVar[str] = "Questões"

    def __init__(self, db_path: str = "output/pipeline_state.db") -> None:
        """Initialize ExcelWriter.

        Args:
            db_path: Path to SQLite database containing questions table
        """
        self.db_path = db_path

    def export_to_excel(self, output_path: str, limit: int | None = None) -> None:
        """Export approved questions to Excel with atomic write.

        Loads all questions with status='approved' from SQLite, creates a DataFrame
        with 26 columns in correct order, and writes to Excel with formatting.
        Uses atomic write pattern (temp file + rename) to prevent corruption.

        Args:
            output_path: Path to output Excel file (.xlsx format)
            limit: Optional maximum number of questions to export (for testing)

        Raises:
            OSError: If database query fails, Excel write fails, or permission denied
            ValueError: If output_path is not .xlsx extension or required columns missing

        Example:
            >>> writer = ExcelWriter()
            >>> writer.export_to_excel("output/final.xlsx")
            INFO: Successfully exported 8432 questions to output/final.xlsx
        """
        # Validate file extension
        if not output_path.endswith(".xlsx"):
            raise ValueError(f"Output path must have .xlsx extension, got: {output_path}")

        output_file = Path(output_path)
        temp_file = Path(str(output_path) + ".tmp")

        try:
            # Create output directory if needed
            self._create_output_directory(output_file)

            # Load approved questions from SQLite
            df = self._load_approved_questions(limit=limit)

            # Handle empty result
            if df.empty:
                logger.warning("No approved questions found - nothing to export")
                # Create empty Excel with headers
                df = pd.DataFrame(columns=self.COLUMN_ORDER)

            # Write to temporary file with formatting
            self._write_excel_with_formatting(df, temp_file)

            # Atomic rename
            os.replace(temp_file, output_file)

            logger.info(f"Successfully exported {len(df)} questions to {output_path}")

        finally:
            # Ensure temp file is cleaned up (runs always, even on exception)
            if temp_file.exists():
                temp_file.unlink()

    def _load_approved_questions(self, limit: int | None = None) -> pd.DataFrame:
        """Load approved questions from SQLite.

        Args:
            limit: Optional maximum number of questions to load

        Returns:
            DataFrame with approved questions

        Raises:
            IOError: If database query fails
        """
        try:
            logger.info("Loading approved questions from SQLite")

            conn = sqlite3.connect(self.db_path)

            # Build query with parameterized LIMIT (prevent SQL injection)
            query = "SELECT * FROM questions WHERE status=? ORDER BY id"
            params = ["approved"]
            if limit:
                query += " LIMIT ?"
                params.append(limit)

            # Load into DataFrame
            df = pd.read_sql_query(query, conn, params=params)
            conn.close()

            logger.info(f"Loaded {len(df)} approved questions")

            # Validate all required columns are present
            missing_columns = set(self.COLUMN_ORDER) - set(df.columns)
            if missing_columns:
                raise ValueError(  # noqa: TRY301
                    f"Database is missing required columns: {sorted(missing_columns)}. "
                    f"Expected all 26 columns from COLUMN_ORDER."
                )

            # Ensure column order and select only needed columns
            # Handle case where DB has extra columns (like id, status)
            df = df[self.COLUMN_ORDER]

            return df

        except ValueError:
            # Re-raise ValueError as-is (validation errors, not IO errors)
            raise
        except sqlite3.Error as e:
            logger.exception("Failed to load questions from SQLite")
            raise OSError(f"Database error: {e}") from e
        except Exception as e:
            logger.exception("Unexpected error loading questions")
            raise OSError(f"Failed to load questions: {e}") from e

    def _create_output_directory(self, output_path: Path) -> None:
        """Create parent directory if doesn't exist.

        Args:
            output_path: Path to output file

        Raises:
            IOError: If directory creation fails
        """
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            if not output_path.parent.exists():
                logger.info(f"Created directory: {output_path.parent}")
        except PermissionError as e:
            logger.exception("Permission denied creating directory")
            raise OSError(f"Cannot create directory {output_path.parent}: {e}") from e

    def _write_excel_with_formatting(self, df: pd.DataFrame, temp_path: Path) -> None:
        """Write DataFrame to Excel and apply formatting.

        Args:
            df: DataFrame to write
            temp_path: Path to temporary file

        Raises:
            IOError: If Excel write or formatting fails
        """
        try:
            logger.info(f"Writing Excel to {temp_path}")

            # Step 1: Write with pandas
            df.to_excel(
                temp_path,
                engine="openpyxl",
                index=False,
                sheet_name=self.DEFAULT_SHEET_NAME,
            )

            # Step 2: Apply formatting with openpyxl
            self._apply_header_formatting(temp_path)

        except Exception as e:
            logger.exception(f"Failed to write Excel to {temp_path}")
            raise OSError(f"Excel write failed: {e}") from e

    def _apply_header_formatting(self, temp_path: Path) -> None:
        """Apply bold formatting to header row.

        Args:
            temp_path: Path to Excel file to format

        Raises:
            IOError: If formatting fails
        """
        try:
            wb = load_workbook(temp_path)
            ws = wb.active

            # Make header row bold
            for cell in ws[1]:  # First row
                cell.font = Font(bold=True)

            wb.save(temp_path)
            wb.close()

        except Exception:
            logger.exception(
                f"Failed to apply header formatting to {temp_path}. "
                "Excel will be created without bold headers (non-critical)."
            )
            # Don't raise - formatting is nice-to-have, not critical
